"""Build judge_score_bootstrap_ci.json from LLM-judge review spreadsheets.

Inputs
------
benchmark_qa_generation_review_*.xlsx under results/, sheet «Оценка_генерации»:
  column 1  — question index in the bench (bench_index);
  column 5  — question text (empty cell = skip row);
  columns 25–28 — four LLM-judge scores: relevance, correctness,
                   faithfulness, completeness; valid range [1, 5].

Output
------
judge_score_bootstrap_ci.json
    One record per (model, bench, mode, metric): point estimate mean
    and 95% confidence interval bounds ci_low / ci_high.

Statistical methodology
-----------------------
Point estimate
    Sample mean — unbiased estimate of the population mean for any sample size.

Confidence interval: nonparametric percentile bootstrap
    Bootstrap over n = 453 (gold) or n = 310 (noise) observations, B = 10,000 replicates.
    95% CI = 2.5th and 97.5th percentiles of the bootstrap distribution of means.

    In code, percentiles are passed in percent units:
        ci = np.percentile(bootstrap_sample_means, [lower_percentile, upper_percentile])

    Percentile bootstrap is used instead of SEM-based intervals (x̄ ± z·σ/√n) because
    judge score distributions are left-skewed (models often get high scores); symmetric
    intervals underestimate uncertainty in the left tail. At n ≈ 450 (gold) and n ≈ 310
    (noise) the approximation is adequate; BCa correction is numerically negligible.

Noise sample construction
    Three independent runs must not be concatenated: rows for the same question across
    runs are repeated measurements, not independent observations. Concatenation would
    inflate n threefold and narrow CIs incorrectly.

    Correct procedure (Anthropic, recommendation #3): for each question present in all
    three runs (matched by bench_index), average each metric separately. Bootstrap uses
    one observation per question (n ≈ 310).

Run: uv run python scripts/build_judge_score_artifacts.py
"""

from __future__ import annotations

import json
import logging
import sys
from collections.abc import Iterator
from pathlib import Path
from typing import Final

import numpy as np
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell

logger = logging.getLogger(__name__)

RowScores = tuple[float, float, float, float]


class JudgeScoreArtifactsBuilder:
    """Reads review Excel and writes judge_score_bootstrap_ci.json.

    All Excel reading goes through _iter_generation_rows — the only place workbooks
    are opened/closed. Cell parsing is isolated in static methods and reused for
    gold and noise without duplication.
    """

    _SHEET: Final[str] = "Оценка_генерации"
    _COL_BENCH_INDEX: Final[int] = 1   # question index — key for noise alignment
    _COL_QUESTION: Final[int] = 5      # non-empty cell = row belongs to a question
    _COL_SCORES: Final[tuple[int, int, int, int]] = (25, 26, 27, 28)  # relevance / correctness / faithfulness / completeness
    _ITER_MAX_COL: Final[int] = 28

    _REVIEW_PREFIX: Final[str] = "benchmark_qa_generation_review_"

    _BOOTSTRAP_JSON: Final[str] = "judge_score_bootstrap_ci.json"

    # Metric order matches columns 25–28 and is fixed across the pipeline.
    _METRIC_KEYS: Final[tuple[str, str, str, str]] = (
        "relevance",
        "correctness",
        "faithfulness",
        "completeness",
    )

    _MODELS: Final[tuple[str, str, str]] = ("gemma_4", "nemotron", "qwen_35")
    _RAG_MODES: Final[tuple[str, str]] = ("baseline", "full")

    # Subdirectory names for three noise runs under results/<model>/results_noise_bench/.
    _NOISE_RUN_DIRS: Final[tuple[str, str, str]] = (
        "first_results_noise_bench",
        "second_results_noise_bench",
        "third_results_noise_bench",
    )

    def __init__(
        self,
        project_root: Path,
        *,
        num_bootstrap_replicates: int = 10_000,
        alpha: float = 0.05,
        random_seed: int = 42,
    ) -> None:
        self._root: Path = project_root.resolve()
        self._results: Path = (self._root / "results").resolve()
        self._num_bootstrap: int = num_bootstrap_replicates
        self._alpha: float = alpha
        self._random_seed: int = random_seed

    def build(self) -> Path:
        """Entry point: rebuild bootstrap JSON and return its path."""
        return self._write_bootstrap_json()

    def _write_bootstrap_json(self) -> Path:
        if not self._results.is_dir():
            logger.error("Missing results directory: %s", self._results)
            sys.exit(1)

        records = self._build_bootstrap_records()
        path = (self._results / self._BOOTSTRAP_JSON).resolve()
        _ = path.write_text(json.dumps(records, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        logger.info("Bootstrap CI: %s records → %s", len(records), path)
        return path

    def _build_bootstrap_records(self) -> list[dict[str, object]]:
        """Outer loop: (model × mode) → gold + noise → four metrics each.

        A single random_generator is passed through all calls so reproducibility
        holds at a fixed random_seed regardless of operation order.
        """
        random_generator = np.random.default_rng(self._random_seed)
        records: list[dict[str, object]] = []

        for model in self._MODELS:
            for rag_mode in self._RAG_MODES:
                gold_workbook_path = (
                    self._results / model / "results_gold_bench" / f"{self._REVIEW_PREFIX}{rag_mode}.xlsx"
                )
                self._append_bootstrap_records(records, model, "gold", rag_mode, self._gold_rows(gold_workbook_path), random_generator)

                noise_averaged_rows = self._noise_rows(model, rag_mode)
                if not noise_averaged_rows:
                    logger.warning(
                        "Noise: empty bench_index intersection for %s / %s — JSON will contain nan",
                        model,
                        rag_mode,
                    )
                self._append_bootstrap_records(records, model, "noise", rag_mode, noise_averaged_rows, random_generator)

        records.sort(key=lambda record: (str(record["model"]), str(record["bench"]), str(record["mode"]), str(record["metric"])))
        return records

    def _gold_rows(self, workbook_path: Path) -> list[RowScores]:
        """All valid rows from one gold file in document order.

        For gold, bench_index is not used for merging runs: one run, one row = one observation.
        """
        valid_rows: list[RowScores] = []
        for row in self._iter_generation_rows(workbook_path):
            if not self._has_question(row):
                continue
            parsed_scores = self._parse_scores(row)
            if parsed_scores is not None:
                valid_rows.append(parsed_scores)
        return valid_rows

    def _noise_rows(self, model: str, rag_mode: str) -> list[RowScores]:
        """Build bootstrap sample for noise: one observation = one question.

        1. Read three noise runs into dicts {bench_index: scores}.
        2. Take key intersection — only questions scored in all three runs.
        3. Average three values per question and metric separately.
        """
        noise_bench_dir = self._results / model / "results_noise_bench"
        per_run_scores = [
            self._rows_indexed_by_bench(noise_bench_dir / run_dir / f"{self._REVIEW_PREFIX}{rag_mode}.xlsx")
            for run_dir in self._NOISE_RUN_DIRS
        ]

        common_bench_indices = set(per_run_scores[0].keys())
        for run_scores_by_index in per_run_scores[1:]:
            common_bench_indices &= set(run_scores_by_index.keys())

        if not common_bench_indices:
            return []

        averaged_rows: list[RowScores] = []
        for question_index in sorted(common_bench_indices):
            # scores_matrix: (3 runs × 4 metrics); mean over runs yields 4 values.
            scores_matrix = np.array(
                [per_run_scores[0][question_index], per_run_scores[1][question_index], per_run_scores[2][question_index]],
                dtype=np.float64,
            )
            metric_means = np.mean(scores_matrix, axis=0)
            averaged_rows.append(
                (
                    float(metric_means.flat[0]),
                    float(metric_means.flat[1]),
                    float(metric_means.flat[2]),
                    float(metric_means.flat[3]),
                )
            )
        return averaged_rows

    def _append_bootstrap_records(
        self,
        output_records: list[dict[str, object]],
        model: str,
        bench: str,
        rag_mode: str,
        rows: list[RowScores],
        random_generator: np.random.Generator,
    ) -> None:
        """Append four records (one per metric) for one (model, bench, mode) combination."""
        scores_matrix = np.asarray(rows, dtype=np.float64) if rows else np.zeros((0, 4), dtype=np.float64)

        for metric_index, metric_name in enumerate(self._METRIC_KEYS):
            metric_column = scores_matrix[:, metric_index] if scores_matrix.size else np.zeros(0, dtype=np.float64)
            output_records.append(
                {
                    "model": model,
                    "bench": bench,
                    "mode": rag_mode,
                    "metric": metric_name,
                    **self._bootstrap_ci(metric_column, random_generator),
                    "n_bootstrap": self._num_bootstrap,
                    "alpha": self._alpha,
                }
            )

    def _rows_indexed_by_bench(self, workbook_path: Path) -> dict[int, RowScores]:
        """Read one noise file into {bench_index: scores}. First row wins on duplicate index."""
        scores_by_bench_index: dict[int, RowScores] = {}
        for row in self._iter_generation_rows(workbook_path):
            if not self._has_question(row):
                continue
            question_index = self._bench_index(row[self._COL_BENCH_INDEX - 1].value)
            if question_index is None:
                continue
            parsed_scores = self._parse_scores(row)
            if parsed_scores is None:
                continue
            if question_index not in scores_by_bench_index:
                scores_by_bench_index[question_index] = parsed_scores
        return scores_by_bench_index

    def _iter_generation_rows(self, workbook_path: Path) -> Iterator[tuple[Cell | MergedCell, ...]]:
        """Yield rows from sheet «Оценка_генерации» starting at row 2 (row 1 is header)."""
        if not workbook_path.is_file():
            return
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            if self._SHEET not in workbook.sheetnames:
                return
            yield from workbook[self._SHEET].iter_rows(min_row=2, max_col=self._ITER_MAX_COL)
        finally:
            workbook.close()

    @staticmethod
    def _bench_index(raw: object) -> int | None:
        """Parse column 1 to integer index > 0; otherwise None."""
        if raw is None or isinstance(raw, bool):
            return None
        if isinstance(raw, int):
            return raw if raw > 0 else None
        if isinstance(raw, float):
            if not raw.is_integer():
                return None
            integer_value = int(raw)
            return integer_value if integer_value > 0 else None
        if isinstance(raw, str):
            stripped = raw.strip()
            if not stripped:
                return None
            try:
                float_value = float(stripped)
            except ValueError:
                return None
            if not float_value.is_integer():
                return None
            integer_value = int(float_value)
            return integer_value if integer_value > 0 else None
        return None

    @staticmethod
    def _has_question(row: tuple[Cell | MergedCell, ...]) -> bool:
        """Row belongs to a question if column 5 is non-empty."""
        cell_value = row[JudgeScoreArtifactsBuilder._COL_QUESTION - 1].value
        return cell_value is not None and str(cell_value).strip() != ""

    def _parse_scores(self, row: tuple[Cell | MergedCell, ...]) -> RowScores | None:
        """Extract four scores from columns 25–28; None if any value is invalid."""
        score_values: list[float] = []
        for col in self._COL_SCORES:
            raw = row[col - 1].value
            if raw is None or isinstance(raw, bool):
                return None
            if isinstance(raw, str):
                stripped = raw.strip()
                if not stripped:
                    return None
                try:
                    score_value = float(stripped)
                except ValueError:
                    return None
            elif isinstance(raw, (int, float)):
                score_value = float(raw)
            else:
                return None
            if not (1.0 <= score_value <= 5.0):
                return None
            score_values.append(score_value)
        return (score_values[0], score_values[1], score_values[2], score_values[3])

    def _bootstrap_ci(self, values: np.ndarray, random_generator: np.random.Generator) -> dict[str, float | int]:
        """Nonparametric percentile bootstrap for one observation vector.

        Returns mean, ci_low, ci_high (two-sided CI at level 1 - alpha) and n.
        Empty vector → nan for all fields.
        """
        n_observations = int(values.size)
        if n_observations == 0:
            return {"mean": float("nan"), "ci_low": float("nan"), "ci_high": float("nan"), "n": 0}

        sample_mean = float(values.mean())

        bootstrap_sample_means = np.empty(self._num_bootstrap, dtype=np.float64)
        for sample_index in range(self._num_bootstrap):
            bootstrap_sample_means[sample_index] = float(
                random_generator.choice(values, size=n_observations, replace=True).mean()
            )

        # Two-sided CI at level (1−α): α/2 in each tail of bootstrap means.
        lower_percentile = 100.0 * self._alpha / 2.0
        upper_percentile = 100.0 * (1.0 - self._alpha / 2.0)
        ci = np.percentile(bootstrap_sample_means, [lower_percentile, upper_percentile])

        return {
            "mean": sample_mean,
            "ci_low": float(ci.flat[0]),
            "ci_high": float(ci.flat[1]),
            "n": n_observations,
        }


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    root = Path(__file__).resolve().parent.parent
    _ = JudgeScoreArtifactsBuilder(root).build()


if __name__ == "__main__":
    main()
